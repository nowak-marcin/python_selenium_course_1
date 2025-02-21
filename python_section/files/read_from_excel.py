from openpyxl import Workbook, load_workbook


wb = load_workbook(filename='demoexcel.xlsx')
sh = wb.active
# wybor innej strony (sheet) niz aktywna:
# sh = wb['demosheet2']

# pobranie konkretnej komorki:

print(sh['A3'].value)
# print(wb['demosheet2']['A2'].value)
print(sh.cell(2, 2).value)

# zliczenie liczby kolumn i wierszy:

row_count = sh.max_row
column_count = sh.max_column

print(row_count, column_count)

# pobranie wybranych wierszy/kolumn:

for i in range(1, row_count+1):              # wszystkie kolejne wiersze od 1 do max
    for j in range(1, column_count+1):       # wszystkie kolejne kolumny od 1 do max
        print(sh.cell(row=i, column=j).value, end=' ')
    print('\n')
