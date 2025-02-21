from openpyxl import Workbook


wb = Workbook()
ws = wb.active

# uzupelnienie konkretnych komorek:
# ws['A1'] = 'club name'
# ws['B1'] = 'city'

# import listy:
testdata = [['club_name', 'city'], ['pogon', 'szczecin'], ['flota', 'swinoujscie'], ['kotwica', 'kolobrzeg']]

for x in testdata:
    ws.append(x)
wb.save('demoexcel.xlsx')

# okreslenie wierszy i kolumn:
# zobacz: libraries_modules -> faker_data
