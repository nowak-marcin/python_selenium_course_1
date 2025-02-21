from faker import Faker
from openpyxl import Workbook


fake_data = Faker('pl_PL')

# basics:
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

# import danych do pliku excel:
wb = Workbook() # inicjalizacja
ws = wb.active  # utworzenie aktywnej strony (worksheet)

# i - liczba wierszy
# j - liczba kolumn
for i in range(1,5):
    for j in range(1,3):
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.address()
wb.save('testData.xlsx')
